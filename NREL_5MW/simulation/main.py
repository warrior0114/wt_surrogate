from openpyxl import load_workbook

from gen_wind import modify_properties

from modify_inflowwind import modify_properties_inflowwind

from modify_ElastoDyn import  modify_properties_ElastoDyn
 

from modify_fst import modify_properties_fst

from run_fst import run_fst

import time
# 加载 Excel 文件
wb = load_workbook("批处理/linshi.xlsx")
sheet = wb.active

for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
    if row_idx >= 1:
        start_time = time.time()
        print("=====" + str(row[0]) + "=====")
        print("=====" + "properties" + "=====")
        print("number", "RandSeed1", "AnalysisTime", "UsableTime", "TurbModel", "IECturbc", "RefHt", "URef", "Yaw","Shear")
        print(row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8],row[9])
        # 生成风文件inp
        RefHt, URef, btsfile_road = modify_properties(row[1], row[2], row[3], row[4], row[5],row[6], row[7],row[8],row[9]) 
        # print(RefHt, URef, btsfile_road)
        # print('\n')
        # 生成inflowwind文件
        # RefHt = 90
        # URef = 12.78
        # btsfile_road = 'wind_63780_90m_12.78mps_TIB_Yaw20_Shear0.14.bts'
        inflowwind_file_road = modify_properties_inflowwind(RefHt, URef, btsfile_road)
        # print(inflowwind_file_road)
        # print('\n')
        # 生成hyd文件
        # WaveMod, WaveHs, WaveTp, hyd_file_road = modify_properties_hyd(row[8], row[9], row[10], row[11])
        # print(WaveMod, WaveHs, WaveTp, hyd_file_road)
        # print('\n')
        #生成偏航角ElastoDyn文件
        ElastoDyn_file_road=modify_properties_ElastoDyn(RefHt, URef, row[8])
        # 生成fst文件
        fst_file_road = modify_properties_fst(RefHt, URef,row[5],row[8],inflowwind_file_road,ElastoDyn_file_road,row[1],row[9])
        # print(fst_file_road)
        # print('\n')
        # 运行fst文件
        run_fst(fst_file_road)
        end_time = time.time()
        print(f"运行时间: {end_time - start_time} 秒")
        print('\n')




    

